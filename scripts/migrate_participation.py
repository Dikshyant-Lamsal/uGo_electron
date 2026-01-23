#!/usr/bin/env python3
"""
Migrate Participation Data from Master_Database to Participations Sheet

This script reads the 'Participation' column from Master_Database and creates
individual participation records in the Participations sheet.
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from datetime import datetime
import re
import os

# Configuration
EXCEL_FILE = 'data/students.xlsx'
MASTER_SHEET = 'Master_Database'
PARTICIPATIONS_SHEET = 'Participations'

def parse_participation_text(participation_text):
    """
    Parse participation text and extract event details.
    
    Common formats:
    - "Workshop on AI - 2023-05-15"
    - "Tech Conference 2023 | Participant | 8 hours"
    - "Volunteer Work, Community Service, 10 hours, 2023-06-20"
    
    Returns list of participation dicts
    """
    if not participation_text or pd.isna(participation_text):
        return []
    
    # Convert to string and clean
    text = str(participation_text).strip()
    if not text or text.lower() in ['none', 'n/a', 'nil', '-']:
        return []
    
    participations = []
    
    # Split by common delimiters (semicolon, newline, double space)
    entries = re.split(r'[;\n]|(?:\s{2,})', text)
    
    for entry in entries:
        entry = entry.strip()
        if not entry or len(entry) < 3:
            continue
        
        participation = {
            'event_name': '',
            'event_date': '',
            'event_type': 'Other',
            'role': 'Participant',
            'hours': 0,
            'notes': entry  # Store original text as notes
        }
        
        # Try to extract event name (everything before date or special chars)
        event_name_match = re.match(r'^([^|,\-:]+)', entry)
        if event_name_match:
            participation['event_name'] = event_name_match.group(1).strip()
        else:
            participation['event_name'] = entry[:50]  # First 50 chars
        
        # Try to extract date (YYYY-MM-DD, DD/MM/YYYY, DD-MM-YYYY)
        date_patterns = [
            r'\b(\d{4}-\d{2}-\d{2})\b',  # 2023-05-15
            r'\b(\d{2}/\d{2}/\d{4})\b',  # 15/05/2023
            r'\b(\d{2}-\d{2}-\d{4})\b',  # 15-05-2023
        ]
        
        for pattern in date_patterns:
            date_match = re.search(pattern, entry)
            if date_match:
                participation['event_date'] = date_match.group(1)
                break
        
        # Try to extract event type
        event_types = ['workshop', 'seminar', 'conference', 'volunteer', 
                      'competition', 'training', 'webinar', 'hackathon']
        entry_lower = entry.lower()
        for event_type in event_types:
            if event_type in entry_lower:
                participation['event_type'] = event_type.capitalize()
                break
        
        # Try to extract role
        roles = ['participant', 'organizer', 'volunteer', 'speaker', 
                'facilitator', 'coordinator', 'member']
        for role in roles:
            if role in entry_lower:
                participation['role'] = role.capitalize()
                break
        
        # Try to extract hours
        hours_match = re.search(r'(\d+)\s*(?:hours?|hrs?|h)\b', entry_lower)
        if hours_match:
            participation['hours'] = int(hours_match.group(1))
        
        participations.append(participation)
    
    return participations


def migrate_participations(excel_file, dry_run=False):
    """
    Main migration function
    
    Args:
        excel_file: Path to Excel file
        dry_run: If True, only print what would be done without modifying the file
    """
    print("=" * 70)
    print("📋 PARTICIPATION MIGRATION SCRIPT")
    print("=" * 70)
    print(f"📁 Excel file: {excel_file}")
    print(f"🔍 Mode: {'DRY RUN (no changes)' if dry_run else 'LIVE (will modify file)'}")
    print()
    
    # Check if file exists
    if not os.path.exists(excel_file):
        print(f"❌ Error: File not found: {excel_file}")
        return
    
    # Read the Excel file
    print("📖 Reading Excel file...")
    try:
        df_master = pd.read_excel(excel_file, sheet_name=MASTER_SHEET)
        print(f"✅ Loaded {len(df_master)} students from {MASTER_SHEET}")
    except Exception as e:
        print(f"❌ Error reading {MASTER_SHEET}: {e}")
        return
    
    # Check if Participations sheet exists
    workbook = load_workbook(excel_file)
    
    if PARTICIPATIONS_SHEET in workbook.sheetnames:
        df_participations = pd.read_excel(excel_file, sheet_name=PARTICIPATIONS_SHEET)
        print(f"📋 Found existing {PARTICIPATIONS_SHEET} sheet with {len(df_participations)} records")
        next_id = df_participations['participation_id'].max() + 1 if len(df_participations) > 0 else 1
    else:
        print(f"📋 Creating new {PARTICIPATIONS_SHEET} sheet")
        df_participations = pd.DataFrame(columns=[
            'participation_id', 'student_id', 'event_name', 'event_date',
            'event_type', 'role', 'hours', 'notes', 'created_at', 'updated_at'
        ])
        next_id = 1
    
    print(f"🔢 Next participation ID will be: {next_id}")
    print()
    
    # Process each student
    new_participations = []
    students_with_data = 0
    total_participations_created = 0
    
    print("🔄 Processing students...")
    print("-" * 70)
    
    for idx, student in df_master.iterrows():
        student_id = student.get('id')
        student_name = student.get('Full_Name', 'Unknown')
        participation_text = student.get('Participation')
        
        # Skip if no student ID
        if pd.isna(student_id):
            continue
        
        # Parse participation text
        participations = parse_participation_text(participation_text)
        
        if participations:
            students_with_data += 1
            print(f"👤 Student #{student_id}: {student_name}")
            print(f"   📝 Original text: {participation_text[:80]}...")
            print(f"   ✅ Found {len(participations)} participation(s)")
            
            for participation in participations:
                new_participation = {
                    'participation_id': next_id,
                    'student_id': int(student_id),
                    'event_name': participation['event_name'],
                    'event_date': participation['event_date'],
                    'event_type': participation['event_type'],
                    'role': participation['role'],
                    'hours': participation['hours'],
                    'notes': participation['notes'],
                    'created_at': datetime.now().isoformat(),
                    'updated_at': datetime.now().isoformat()
                }
                
                print(f"      ➜ [{next_id}] {participation['event_name'][:40]} | "
                      f"{participation['event_type']} | {participation['hours']}h")
                
                new_participations.append(new_participation)
                next_id += 1
                total_participations_created += 1
            
            print()
    
    # Summary
    print("=" * 70)
    print("📊 MIGRATION SUMMARY")
    print("=" * 70)
    print(f"👥 Total students processed: {len(df_master)}")
    print(f"✅ Students with participation data: {students_with_data}")
    print(f"📋 New participation records created: {total_participations_created}")
    print(f"📝 Existing participation records: {len(df_participations)}")
    print(f"📊 Total after migration: {len(df_participations) + total_participations_created}")
    print()
    
    # Save to Excel if not dry run
    if not dry_run and new_participations:
        print("💾 Saving to Excel...")
        
        # Combine existing and new participations
        df_new = pd.DataFrame(new_participations)
        df_combined = pd.concat([df_participations, df_new], ignore_index=True)
        
        try:
            # Write back to Excel
            with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', 
                               if_sheet_exists='replace') as writer:
                df_combined.to_excel(writer, sheet_name=PARTICIPATIONS_SHEET, index=False)
            
            print(f"✅ Successfully saved {len(df_combined)} records to {PARTICIPATIONS_SHEET}")
            print(f"📁 File updated: {excel_file}")
        except Exception as e:
            print(f"❌ Error saving to Excel: {e}")
            print("\n📋 Here's the data that would have been saved:")
            print(df_new.to_string())
    elif dry_run:
        print("ℹ️  DRY RUN - No changes made to the file")
        if new_participations:
            print("\n📋 Preview of data that would be created:")
            df_preview = pd.DataFrame(new_participations)
            print(df_preview.to_string(max_rows=10))
    else:
        print("ℹ️  No new participations to add")
    
    print()
    print("✅ Migration complete!")


def main():
    """Main entry point"""
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Migrate participation data from Master_Database to Participations sheet'
    )
    parser.add_argument(
        '--file', '-f',
        default=EXCEL_FILE,
        help=f'Path to Excel file (default: {EXCEL_FILE})'
    )
    parser.add_argument(
        '--dry-run', '-d',
        action='store_true',
        help='Perform a dry run without modifying the file'
    )
    parser.add_argument(
        '--test',
        action='store_true',
        help='Test participation text parsing with sample data'
    )
    
    args = parser.parse_args()
    
    if args.test:
        # Test the parser with sample data
        print("🧪 Testing participation text parser...")
        print("=" * 70)
        
        test_cases = [
            "Workshop on AI - 2023-05-15",
            "Tech Conference 2023 | Participant | 8 hours",
            "Volunteer Work, Community Service, 10 hours, 2023-06-20",
            "Seminar; Training session; Hackathon 2023",
            "Speaker at Workshop - 15/05/2023 - 3 hours",
            "",
            "N/A",
            "Multiple events:\nWorkshop A - 2023-01-10\nSeminar B - 2023-02-15"
        ]
        
        for i, test_text in enumerate(test_cases, 1):
            print(f"\n📝 Test Case {i}:")
            print(f"   Input: {repr(test_text)}")
            results = parse_participation_text(test_text)
            if results:
                for participation in results:
                    print(f"   ✅ Event: {participation['event_name']}")
                    print(f"      Type: {participation['event_type']}, "
                          f"Role: {participation['role']}, "
                          f"Hours: {participation['hours']}")
                    if participation['event_date']:
                        print(f"      Date: {participation['event_date']}")
            else:
                print("   ❌ No participations parsed")
        
        print("\n" + "=" * 70)
        return
    
    # Run migration
    migrate_participations(args.file, dry_run=args.dry_run)


if __name__ == '__main__':
    main()